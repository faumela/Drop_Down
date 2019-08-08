import openpyxl

#load workbook

wk = openpyxl.load_workbook("C:\\Users\\ANJUM\\Drop_Down\\Drivers\\Tests.xlsx")

print(wk.sheetnames)

print("Active Sheet is " + wk.active.title)

# Create object of any sheet on which you want to work

sh = wk ['Tests']
print(sh.title)

# Fetch the value
print(sh['A4'].value)
print(sh['B3'].value)
print(sh['C2'].value)

# Creating Cell Object


cl = sh.cell(3,2)
print(cl.value)

# Find Rows having Data
rows = sh.max_row
columns = sh.max_column

print("Total Rows are - " + str(rows))
print("Total columns are - " + str(columns))

for i in range(1,rows+1):
    for j in range(1,columns+1):
         c= sh.cell(i,j)
         print(c.value)

for r in sh['A1' : 'C4']:
    for c in r:
        print(c.value)















